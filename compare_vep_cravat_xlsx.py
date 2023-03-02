#!/usr/bin/env python3

import pandas as pd
import numpy as np
import argparse
import logging
import os
import sys
import yaml
import openpyxl


def main():
    logger = logging.getLogger()

    parser = argparse.ArgumentParser(
        description="Compare VEP vs OpenCRAVAT excel report",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument(
        "-v", "--vep-path", required=True,
        help="Path to VEP annotated Excel file",
    )
    parser.add_argument(
        "-c", "--cravat-path", required=True,
        help="Path to OpenCRAVAT annotated Excel file",
    )
    parser.add_argument(
        "-o", "--output", required=True,
        help="Path to output comparison file",
    )
    parser.add_argument(
        "-d", "--drop", action='store_true',
        help="Drop equal values",
    )
    parser.add_argument(
        "-m", "--merge", choices=['inner', 'outer', 'left', 'right'], default='inner',
        help="Apply chosen merge method",
    )
    args = parser.parse_args()

    vep_path = os.path.abspath(args.vep_path)
    cravat_path = os.path.abspath(args.cravat_path)
    output_path = os.path.abspath(args.output)

    if not all([
        check_file_existence(vep_path, logger),
        check_file_existence(cravat_path, logger),
        check_dir_existence(os.path.dirname(output_path), logger)
    ]):
        sys.exit(1)


    try:
        # Reading input xlsx files
        logger.warning('\nReading xlsx ...')
        vep_df = xlsx2df(
            file_name=vep_path,
            sheet_name='data',
            hyperlink_function_columns=['POS', 'GCards', 'CSQ_Existing_variation', 'GNOMAD', 'OMIM_LINK'],
            value_function_columns=['CSQ_CADD_PHRED']
        )
        cravat_df = xlsx2df(
            file_name=cravat_path,
            sheet_name='Variant',
            hyperlink_function_columns=['POS', 'GCards', 'Existing_variation', 'GNOMAD', 'OMIM_LINK'],
            value_function_columns=['cadd_phred']
        )


        # Merging vep and cravat pandas.DataFrames
        logger.warning('Merging ...')
        merge_on_cols = [
            'CHROM',
            'POS',
            'REF',
            'ALT',
            'GCards'  
        ]

        merged_df = pd.merge(
            vep_df,
            cravat_df,
            how=args.merge,
            left_on=merge_on_cols,
            right_on=merge_on_cols,
            suffixes=('_vep', '_cravat')
        )

        # Set comparison column pairs and custom dtypes for normalization style
        comparison_cols = {
            ('GT_vep', 'GT_cravat'): 'str',
            ('VAF_vep', 'VAF_cravat'): 'float',
            ('CSQ_Existing_variation', 'Existing_variation'): 'str',
            ('GeneRegion_vep', 'GeneRegion_cravat'): 'str_list',
            ('VariantType_vep', 'VariantType_cravat'): 'str',
            ('CSQ_Consequence', 'Consequence'): 'str_list',
            ('CSQ_Feature', 'Feature'): 'str_list',
            ('CSQ_SpliceRegion', 'SpliceRegion'): 'str_list',
            ('CSQ_IMPACT', 'IMPACT'): 'str_list',
            ('CSQ_EXON', 'EXON'): 'str_list',
            ('CSQ_HGVSc', 'HGVSc'): 'str_list',
            ('CSQ_HGVSp', 'HGVSp'): 'str_list_hgvsp',
            ('CSQ_HGVSg', 'HGVSg'): 'str',
            ('CSQ_BIOTYPE', 'BIOTYPE'): 'str_list',
            ('MGI_vep', 'MGI_cravat'): 'str_list_3',
            ('gnomADg3_MAX_vep', 'gnomADg3_MAX_cravat'): 'float',
            ('gnomADg3_NHOM_MAX', 'gnomADg3_nhomalt'): 'float',
            ('CSQ_CLIN_SIG', 'ClinSig'): 'str_clinsig',
            ('Panel_vep', 'Panel_cravat'): 'str',
            ('CSQ_PUBMED', 'Pubmed'): 'str_list_pubmed',
            ('OMIM_LINK_vep', 'OMIM_LINK_cravat'): 'float_omim',
            ('OMIM_PHE_HPO_vep', 'OMIM_PHE_HPO_cravat'): 'str_list_2',
            ('HPO_Term_Name', 'HPO_Term'): 'str_list_2',
            ('HPO_inheritance', 'HPO_Inheritance'): 'str',
            ('CSQ_CADD_PHRED', 'cadd_phred'): 'float',
            ('CSQ_SpliceAI_cutoff', 'spliceai_cutoff'): 'str',
            ('CSQ_ada_score', 'ada_score'): 'float',
            ('CSQ_SIFT', 'sift_pred'): 'str',
            ('CSQ_PolyPhen', 'polyphen_hvar_pred'): 'str'
        }
        ordered_columns = merge_on_cols + list(sum(comparison_cols.keys(), ()))
    
        merged_ordered_df = merged_df[ordered_columns]


        # Dropping equal values of each comparison column pair (optional)
        if args.drop:
            logger.warning('Dropping equal ...')
            merged_ordered_df = merged_ordered_df \
                .apply(drop_equal, axis=1, result_type=None, comparison_cols=comparison_cols)


        # Colorizing values of each comparison column pair
        logger.warning('Colorizing ...')
        merged_ordered_df = merged_ordered_df.style \
            .set_properties(**{'background-color': '#fbfaea'}, subset=merged_ordered_df.columns[5::4]) \
            .set_properties(**{'background-color': '#fbfaea'}, subset=merged_ordered_df.columns[6::4]) \
            .apply(colorize_values, axis=1, comparison_cols=comparison_cols)


        # Writing output xlsx
        logger.warning('Writing xlsx ...')
        merged_ordered_df.to_excel(output_path, index=False, freeze_panes=(1, 5))

        logger.warning(f'Done: {output_path}\n')

    except Exception as e:
        logger.critical(e)


# ========= pandas ===========

def xlsx2df(
    file_name: str,
    sheet_name: str,
    hyperlink_function_columns: list[str],
    value_function_columns: list[str],
    row_header: int = 1
) -> pd.DataFrame:

    df = pd.read_excel(file_name, sheet_name)
    ws = openpyxl.load_workbook(file_name)[sheet_name]
    for column in hyperlink_function_columns:
        row_offset = row_header + 1
        column_index = list(df.columns).index(column) + 1
        df[column] = [
            extract_hyperlink_function_value(ws.cell(row=row_offset + i, column=column_index))
            for i in range(len(df[column]))
        ]
    for column in value_function_columns:
        row_offset = row_header + 1
        column_index = list(df.columns).index(column) + 1
        df[column] = [
            extract_value_function_value(ws.cell(row=row_offset + i, column=column_index))
            for i in range(len(df[column]))
        ] 
    return df


def extract_hyperlink_function_value(cell):
    try:
        return cell.value.split(',')[-1].lstrip().lstrip('"').rstrip('")')
    except:
        return None


def extract_value_function_value(cell):
    try:
        left_str_idx = cell.value.find('(')
        right_str_idx = cell.value.find(')')
        extracted_number = float(cell.value[left_str_idx+2:right_str_idx-1].replace(',', '.'))
        return extracted_number
    except:
        return None


def drop_equal(series: pd.Series, comparison_cols: dict) -> pd.Series:
    for pair_cols, dtype in comparison_cols.items():
        pair = normalize_pair((series[pair_cols[0]], series[pair_cols[1]]), dtype)
        if any([
            pair[0] == pair[1],
            ('list' in dtype and is_equal_set(pair[0], pair[1])),
            (dtype == 'float' and np.round(pair[0], 2) == np.round(pair[1], 2))
        ]): series[pair_cols[0]], series[pair_cols[1]] = np.nan, np.nan
    return series


def colorize_values(series: pd.Series, comparison_cols: dict) -> list:
    colors = ['color: black'] * 5
    for pair_cols, dtype in comparison_cols.items():
        if pair_cols[0] not in series.index:
            continue
        pair = normalize_pair((series[pair_cols[0]], series[pair_cols[1]]), dtype)
        if pair[0] == pair[1]:
            colors += ['color: gray'] * 2
        elif 'list' in dtype:
            if is_equal_set(pair[0], pair[1]): # grey
                colors += ['color: grey'] * 2
            elif is_subset(pair[0], pair[1])['of_cravat']: # green
                colors += ['color: #009415'] * 2
            elif is_subset(pair[0], pair[1])['of_vep']: # red
                colors += ['color: #f05300'] * 2
            else: # black
                colors += ['color: black'] * 2 
        elif dtype == 'float':
            if np.round(pair[0], 2) == np.round(pair[1], 2): # gray
                colors += ['color: gray'] * 2
            else: # black
                colors += ['color: black'] * 2
        else: # black
            colors += ['color: black'] * 2
    return colors


def normalize_pair(pair: tuple, dtype: str) -> tuple:
    if not pair[0] or not pair[1]:
        return pair
    if pd.isna(pair[0]) or pd.isna(pair[1]):
        return pair
    if dtype == 'str_list':
        pair = normalize_str_list(pair)
    elif dtype == 'str_list_2':
        pair = normalize_str_list(pair, ', ', ', ')
    elif dtype == 'str_list_3':
        pair = normalize_str_list(pair, '|', ';')
    elif dtype == 'str_list_feat':
        pair = normalize_feature(pair)
    elif dtype == 'str_list_hgvsp':
        pair = normalize_hgvsp(pair)
    elif dtype == 'str_list_pubmed':
        pair = normalize_pubmed(pair)
    elif dtype == 'str_clinsig':
        pair = normalize_clinsig(pair)
    elif dtype == 'float_omim':
        pair = normalize_omimlink(pair)
    return pair


def normalize_str_list(pair: tuple, delimiter_vep=',', delimiter_cravat=',') -> tuple:
    str1 = ','.join(filter(None, pair[0].split(delimiter_vep)))
    str2 = ','.join(filter(None, pair[1].split(delimiter_cravat)))
    return (str1, str2)


def normalize_feature(pair: tuple) -> tuple:
    str1 = ','.join([tr for tr in pair[0].split(',') if tr.startswith('ENST')])
    str2 = pair[1]
    return (str1, str2)


def normalize_clinsig(pair: tuple) -> tuple:
    str1 = pair[0].lower()
    str2 = pair[1].lower()
    return (str1, str2)


def normalize_omimlink(pair: tuple) -> tuple:
    omimlink1 = yaml.safe_load(pair[0])
    omimlink1 = int(omimlink1[0])
    omimlink2 = int(pair[1])
    return (omimlink1, omimlink2)


def normalize_hgvsp(pair: tuple) -> tuple:
    hgvsp1 = set(filter(None, [hgvsp.split(':')[-1] for hgvsp in str(pair[0]).split(',')]))
    hgvsp2 = set(filter(None, [hgvsp.split(':')[-1] for hgvsp in str(pair[1]).split(',')]))
    return (hgvsp1, hgvsp2)


def normalize_pubmed(pair: tuple) -> tuple:
    pubmed_vep = ','.join(filter(None, pair[0].split('&')))
    pubmed_cravat = ','.join(filter(None, pair[1].split(', ')))
    return (pubmed_vep, pubmed_cravat)


def is_equal_set(str1: str, str2: str) -> bool:
    set1 = set(filter(None, str(str1).split(',')))
    set2 = set(filter(None, str(str2).split(',')))
    return set1 == set2


def is_subset(str1: str, str2: str) -> dict:
    set1 = set(filter(None, str(str1).split(',')))
    set2 = set(filter(None, str(str2).split(',')))
    return {
        'of_cravat': set1.issubset(set2),
        'of_vep': set2.issubset(set1)
    }


# ========= utils ===========

def check_file_existence(path: str, logger: logging.Logger) -> bool:
    if not os.path.isfile(path):
        logger.critical(f'No such file: "{path}"')
        return False
    return True


def check_dir_existence(path: str, logger: logging.Logger) -> bool:
    if not os.path.isdir(path):
        logger.critical(f'No such directory: "{path}"')
        return False
    return True


# ========= ====== ===========

if __name__ == '__main__':
    main()
